import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import openpyxl
from datetime import datetime, timedelta
import os
import win32com.client as win32
import numpy as np
from datetime import timedelta
import pandas as pd
import statsmodels.api as sm

# Função para enviar um email
def enviar_email(para, assunto, corpo):
    outlook = win32.Dispatch('Outlook.Application')
    mail = outlook.CreateItem(0)
    mail.To = para
    mail.Subject = assunto
    mail.Body = corpo
    mail.Send()

# Suas credenciais do Twilio
account_sid ='AC0a74e009e3115f542403f0c671cb0388_'
auth_token = 'da6b9d0d7bec1e9afa50127dd6ddcf7e'

# Função para enviar uma mensagem SMS
def enviar_sms(numero_telefone, mensagem):
    client = Client(account_sid, auth_token)

    try:
        message = client.messages.create(
            body=mensagem,
            from_='+19204623603',
            to=numero_telefone
        )
        return message.sid
    except Exception as e:
        return str(e)


# Function to calculate and display the forecast for the current month
# Function to calculate and display the exponential forecast for the current month
def forecast_exponential():
    global forecast_fabrico, forecast_granulacao,forecast_kwh_elect, forecast_kwh_term,forecast_kwh_ton, forecast_kwht_ton
    
    global avg_granulacao
    global avg_kwh_electrico
    global avg_kwh_termico
    global avg_kw_ton
    global avg_kwt_ton
    global previsao_mensal
    sum_fabrico = 0
    sum_granulacao = 0
    sum_kwh_electrico = 0
    sum_kwh_termico = 0
    sum_kw_ton = 0
    sum_kwt_ton = 0
    # Load the data from Sheet2
    
    excel_filename = "dados.xlsx"

    df = pd.read_excel(excel_filename, sheet_name="Sheet1")

    # Calculate the average of the last 22 records for "Fabrico"
    last_22_records = df.tail(22)  # Get the last 22 records
    average_fabrico = last_22_records["Fabrico"].mean()
    avg_granulacao  = last_22_records["Granulação"].mean()*22
    avg_kwh_electrico = last_22_records["Kwh electrico"].mean()*22
    avg_kwh_termico = last_22_records["Kwh termico"].mean()*22
    avg_kw_ton = last_22_records["Kwh/ton"].mean()*22
    avg_kwt_ton = last_22_records["Kwh/ton_termico"].mean()*22

    
    # Calculate the monthly forecast
    previsao_mensal = average_fabrico * 22

    # Load the Excel workbook using openpyxl
    workbook = openpyxl.load_workbook(excel_filename)

    # Create or load "Sheet3" for storing forecasted values


    avg_granulacao


    
    if "Sheet3" not in workbook.sheetnames:
        workbook.create_sheet("Sheet3")

    # Select "Sheet3"
    sheet3 = workbook["Sheet3"]

    # Append the monthly forecast to "Sheet3"
    sheet3.append(["previsao_mensal", previsao_mensal])

    # Save the workbook
    workbook.save(excel_filename)

    
    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet2 = workbook["Sheet2"]

        # Calculate the sum of previous months for each category
        sum_fabrico = 0
        sum_granulacao = 0
        sum_kwh_electrico = 0
        sum_kwh_termico = 0
        sum_kw_ton = 0
        sum_kwt_ton = 0

        for row in sheet2.iter_rows(min_row=2, max_row=sheet2.max_row-1, values_only=True):
            sum_fabrico += row[2]  # Assuming 'Fabrico' is in the third column
            sum_granulacao += row[3]  # Assuming 'Granulação' is in the fourth column
            sum_kwh_electrico += row[4]  # Assuming 'Kwh electrico' is in the fifth column
            sum_kwh_termico += row[5]  # Assuming 'Kwh termico' is in the sixth column
            sum_kw_ton += row[6]  # Assuming 'Kwh/ton' is in the seventh column
            sum_kwt_ton += row[7]  # Assuming 'Kwh/ton_termico' is in the eighth column


           
        sheet4 = workbook["Sheet4"]
        data4 = ["Average Values", average_fabrico, avg_granulacao, avg_kwh_electrico, avg_kwh_termico, avg_kw_ton, avg_kwt_ton, previsao_mensal]
        sheet4.append(data4)
        workbook.save(excel_filename)

        # Display the average values
        messagebox.showinfo("Previsão Mensal para o Mês Atual",
                            f"Previsão Mensal para o Mês Atual:\n"
                            f"Fabrico previsão mensal {previsao_mensal :.2f}\n"
                            f"Granulação: {avg_granulacao:.2f}\n"
                            f"Kwh elétrico: {avg_kwh_electrico:.2f}\n"
                            f"Kwh térmico: {avg_kwh_termico:.2f}\n"
                            
                            f"Kwh/ton: {avg_kw_ton:.2f}\n"
                            f"Kwh/ton: {avg_kwt_ton:.2f}\n")

        # Create or load "Sheet3" for storing forecasted values
        if "Sheet3" not in workbook.sheetnames:
            workbook.create_sheet("Sheet3")

        sheet3 = workbook["Sheet3"]

        # Convert year and month columns into a datetime index
        df = pd.read_excel(excel_filename, sheet_name="Sheet2")
        df['Data'] = pd.to_datetime(df['Ano'].astype(str) + '-' + df['Mês'].astype(str), format='%Y-%m')
        df.set_index('Data', inplace=True)

        # Perform exponential forecasting using ETS (Exponential Smoothing)
        # You may need to adjust the alpha parameter based on your data
        alpha = 0.2
        alpha = 0.2
        use_boxcox = True  # Defina o valor de use_boxcox aqui
        model = sm.tsa.ExponentialSmoothing(df['Soma Fabrico'], seasonal='add', seasonal_periods=12, use_boxcox=use_boxcox)
        results = model.fit(smoothing_level=alpha)
        forecast_fabrico = results.forecast(steps=1).iloc[0]
        # Write the forecasted values to "Sheet3"
        sheet3.append(["Fabrico", forecast_fabrico])

        model = sm.tsa.ExponentialSmoothing(df['Soma Granulação'], seasonal='add', seasonal_periods=12, use_boxcox=use_boxcox)
        results = model.fit(smoothing_level=alpha)
        forecast_granulacao = results.forecast(steps=1).iloc[0]
         
        
        sheet3.append(["Soma Granulação", forecast_granulacao])

     ##############################################################################
        
        model = sm.tsa.ExponentialSmoothing(df['Soma Kwh elétrico'], seasonal='add', seasonal_periods=12, use_boxcox=use_boxcox)
        results = model.fit(smoothing_level=alpha)
        forecast_kwh_elect = results.forecast(steps=1).iloc[0]


        # Write the forecasted values to "Sheet3"
        sheet3.append(["Soma Kwh elétrico", forecast_kwh_elect])
     ##############################################################################
        
        model = sm.tsa.ExponentialSmoothing(df['Soma Kwh térmico'], seasonal='add', seasonal_periods=12, use_boxcox=use_boxcox)
        results = model.fit(smoothing_level=alpha)
        forecast_kwh_term = results.forecast(steps=1).iloc[0]


        # Write the forecasted values to "Sheet3"
        sheet3.append(["Soma Kwh térmico", forecast_kwh_term])
     ##############################################################################
        
        model = sm.tsa.ExponentialSmoothing(df['Soma Kwh/ton'], seasonal='add', seasonal_periods=12, use_boxcox=use_boxcox)
        results = model.fit(smoothing_level=alpha)
        forecast_kwh_ton = results.forecast(steps=1).iloc[0]


        # Write the forecasted values to "Sheet3"
        sheet3.append(["Soma Kwh/ton", forecast_kwh_ton])
     ##############################################################################
        
        model = sm.tsa.ExponentialSmoothing(df['Soma Kwh ton termico'], seasonal='add', seasonal_periods=12, use_boxcox=use_boxcox)
        results = model.fit(smoothing_level=alpha)
        forecast_kwht_ton = results.forecast(steps=1).iloc[0]
       

        # Write the forecasted values to "Sheet3"
        sheet3.append(["Soma Kwh/ton termico", forecast_kwht_ton])                

        # Save the workbook with the forecasted values
        workbook.save(excel_filename)

        # Display the forecasted values
        messagebox.showinfo("Previsão Exponencial para o Mês Atual",
                            f"Previsão Exponencial para o Mês Atual:\n"
                            f"Fabrico: {forecast_fabrico:.2f}\n"
                            f"Granulação :{forecast_granulacao:.2f}\n"
                            f"Kwh electrico:{forecast_kwh_elect:.2f}\n"
                            f"Kwh termico :{forecast_kwh_term:.2f}\n"
                            f"Fabrico previsão mensal {previsao_mensal :.2f}\n"
                            f"Kwh electrico /ton :{forecast_kwh_ton:.2f}\n"
                            f"Kwh termico /ton :{forecast_kwht_ton:.2f}\n")
        

    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao calcular a previsão exponencial foi feita uma previsão baseada nas médias: {str(e)}")
        
    finally :
        return  avg_granulacao,  avg_kwh_electrico,  avg_kwh_termico,  avg_kw_ton,  avg_kwt_ton, previsao_mensal

# Função para salvar os dados
def salvar_dados():

    
    
    mail=entries["Mail"].get()
    fabrico = entries["Fabrico"].get()
    granulacao = entries["Granulação"].get()
    Kwh_electrico = entries["Kwh electrico"].get()
    Kwh_termico = entries["Kwh termico"].get()
    kw_ton=float(Kwh_electrico)/float(fabrico)
    kwt_ton=float(Kwh_termico)/float(granulacao)

    data_=entries["Data"].get()
    
    #data_ = datetime.now().strftime("%Y-%m-%d")

    mensagem = []
   
    try :
        #mensagem = "Segue em email a informação dos consumos energéticos:  " + data_ + "\n A produção em toneladas foi de "+ fabrico+ "\n Em termos de toneladas em granulação foi " + granulacao +"\n Em termos de de energia electrica consumida KWh foi" +str(Kwh_electrico) +"\n Em termos de energia termica KWh foi " + Kwh_termico +"\n Em em termos de KWh por tonelada foi "+str(kw_ton)+"\n Em termos de KWh termico foi  "+str(kwt_ton)+" \n Em termos de previsão de fabrico " + str(forecast_fabrico) +";\n Em termos de previsão toneladas na Granulação " + str(forecast_granulacao)+";\n Em temos de  previsão de consumo electrico " +str(forecast_kwh_elect)+";\n Em temos de  previsão de consumo termico " +str(forecast_kwh_term) +";\n Em termos de  previsão KWH/TON electrico  " +str(forecast_kwh_ton)+";\n Em termos de  previsão KWH/TON termico " + str(forecast_kwht_ton)
        mensagem = (f"Data de fabrico {data_}\n"
            f"Toneladas dosificada {fabrico}\n"
            f"Toneladas Granuladas {granulacao}\n"
            f"Consumo EE  KWh {Kwh_electrico}\n"
            f"Consumo ET KWh {Kwh_termico}\n"
            f"KWh/T EE   {kw_ton:.2f}\n"
            f"KWh/T ET  {kwt_ton:.2f}\n"f"Fabrico previsão mensal  {previsao_mensal :.2f}\n"
            f"Previsão KWh/T EE {forecast_kwh_ton:.2f}\n"
            f"Previsão KWh/T ET {forecast_kwht_ton:.2f}"
            )     
        
    except Exception as e: 
        
        #mensagem = "Segue em email a informação dos consumos energéticos:  " (f"({data_:2f)}")+";\n A produção em toneladas foi de " + fabrico +";\n Em temos de toneladas em granulação foi " +granulacao + ";\n Em termos de de energia electrica consumida KWh foi" +str(Kwh_electrico) + "\n Em termos de energia termica KWh foi " + Kwh_termico +"\n Em em termos de KWh por tonelada foi "+str(kw_ton)+"\n Em termos de KWh termico foi  "+str(kwt_ton)+" \n Em termos de previsão de fabrico " + str(avg_fabrico) + ";\n Em temos de previsão toneladas na Granulação " + str(avg_granulacao)+";\n Em temos de  previsão de consumo electrico " + str(avg_kwh_electrico)+";\n Em temos de  previsão de consumo termico " + str(avg_kwh_termico) +";\n Em temos de  previsão KWH/TON electrico  " + str(avg_kw_ton)+";\n Em temos de  previsão KWH/TON termico " + str(avg_kwt_ton)
        mensagem = (f"Data de fabrico {data_}\n"
            f"Toneladas dosificadas: {fabrico}\n"
            f"Toneladas Granuladas: {granulacao}\n"
            f"Consumo EE KWh: {Kwh_electrico}\n"
            f"Consumo ET KWh: {Kwh_termico}\n"
            f"KWh/T EE: {kw_ton:.2f}\n"
            f"KWh/T ET: {kwt_ton:.2f}\n"f"Fabrico previsão mensal {previsao_mensal :.2f}\n"
            f"Previsão KWh/T EE: {avg_kw_ton:.2f}\n"
            f"Previsão KWh/T ET: {avg_kwt_ton:.2f}"
        )


        
    finally:
        try:
            enviar_email(mail, "Informação de consumo energético", mensagem)
            
        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao enviar o email - Por favor verifique o mail")

            
    # Obter a data e hora atual

    # Cria um novo arquivo Excel
    excel_filename = "dados.xlsx"
    if not os.path.isfile(excel_filename):
        workbook = openpyxl.Workbook()
        workbook.create_sheet("Sheet1")
        workbook.create_sheet("Sheet2")
        sheet1 = workbook["Sheet1"]
        sheet2 = workbook["Sheet2"]

        # Define os cabeçalhos das colunas na Sheet1
        headers = ["Data", "Fabrico", "Granulação", "Kwh electrico", "Kwh termico",'Kwh/ton', 'Kwh/ton_termico' ]
        for col_num, header in enumerate(headers, 1):
            sheet1.cell(row=1, column=col_num, value=header)

        # Define os cabeçalhos das colunas na Sheet2
        headers = ["Mês", "Soma Fabrico", "Soma Granulação", "Soma Kwh elétrico", "Soma Kwh térmico", 'Soma Kwh/ton', ' Soma Kwh ton termico' ]
        for col_num, header in enumerate(headers, 1):
            sheet2.cell(row=1, column=col_num, value=header)
    else:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet1 = workbook["Sheet1"]
        sheet2 = workbook["Sheet2"]
        
    # Escreve os dados nas linhas subsequentes na Sheet1
    data = [data_, fabrico, granulacao, Kwh_electrico, Kwh_termico,kw_ton,kwt_ton  ]
    
    sheet1.append(data)

    # Salva o arquivo Excel com um nome único
    workbook.save(excel_filename)

    # Calcular a soma dos valores de "Fabrico," "Granulação," "Kwh elétrico," e "Kwh térmico" por mês na Sheet2
    df = pd.read_excel(excel_filename, sheet_name="Sheet1")




    # Converta a coluna "Data" para o tipo de data
    df['Data'] = pd.to_datetime(df['Data'])

    # Crie uma nova coluna "Mês" para armazenar o mês de cada registro
    # Crie novas colunas "Ano" e "Mês" para armazenar o ano e o mês de cada registro
    df['Ano'] = df['Data'].dt.year
    df['Mês'] = df['Data'].dt.month

    # Use o pandas para calcular as somas por ano e mês
    soma_por_ano_mes = df.groupby(['Ano', 'Mês']).agg({
        'Fabrico': 'sum',
        'Granulação': 'sum',
        'Kwh electrico': 'sum',
        'Kwh termico': 'sum',
        'Kwh/ton': 'sum',
        'Kwh/ton_termico': 'sum'
    }).reset_index()

    # Atualize o arquivo Excel com as somas por ano e mês na Sheet2
    sheet2.cell(row=1, column=1, value="Ano")
    sheet2.cell(row=1, column=2, value="Mês")
    sheet2.cell(row=1, column=3, value="Soma Fabrico")
    sheet2.cell(row=1, column=4, value="Soma Granulação")
    sheet2.cell(row=1, column=5, value="Soma Kwh elétrico")
    sheet2.cell(row=1, column=6, value="Soma Kwh térmico")
    sheet2.cell(row=1, column=7, value="Soma Kwh/ton")
    sheet2.cell(row=1, column=8, value="Soma Kwh ton termico")
    for i, row in soma_por_ano_mes.iterrows():
        sheet2.cell(row=i+2, column=1, value=row['Ano'])
        sheet2.cell(row=i+2, column=2, value=row['Mês'])
        sheet2.cell(row=i+2, column=3, value=row['Fabrico'])
        sheet2.cell(row=i+2, column=4, value=row['Granulação'])
        sheet2.cell(row=i+2, column=5, value=row['Kwh electrico'])
        sheet2.cell(row=i+2, column=6, value=row['Kwh termico'])
        sheet2.cell(row=i+2, column=7, value=row['Kwh/ton'])
        sheet2.cell(row=i+2, column=8, value=row['Kwh/ton_termico'])
   
 

 

    # Salve o arquivo Excel atualizado
    workbook.save(excel_filename)

    # Exibe uma mensagem de confirmação
    messagebox.showinfo("Sucesso", f"Os dados foram salvos em {excel_filename}")

    # Atualiza a área de exibição
    atualizar_exibicao_dados(10)

def atualizar_exibicao_dados(n_linhas):
    global exibicao_dados  # Declara exibicao_dados como variável global

    # Lê os dados da Sheet1 do arquivo Excel e exibe-os na área de exibição
    excel_filename = "dados.xlsx"
    try:
        workbook = openpyxl.load_workbook(excel_filename)
        sheet1 = workbook["Sheet1"]
        data = list(sheet1.iter_rows(values_only=True))[-n_linhas:]  # Pega as últimas n_linhas linhas de dados da Sheet1

        exibicao_dados.config(state=tk.NORMAL)
        exibicao_dados.delete(1.0, tk.END)  # Limpa o texto atual na área de exibição
        exibicao_dados.insert(tk.END, "Últimas 5 linhas de dados na Sheet1:\n\n")

        for row in data:
            for header, value in zip(sheet1[1], row):
                exibicao_dados.insert(tk.END, f"{header.value}: {value}\n")
            exibicao_dados.insert(tk.END, "\n")

        exibicao_dados.config(state=tk.DISABLED)
    except Exception as e:
        exibicao_dados.config(state=tk.NORMAL)
        exibicao_dados.delete(1.0, tk.END)
        exibicao_dados.insert(tk.END, f"Erro ao carregar os dados da Sheet1: {str(e)}")
        exibicao_dados.config(state=tk.DISABLED)

# Criação da janela principal
root = tk.Tk()
root.title("Formulário de Entrada de Dados")
root.geometry("1000x1000")

# Criação dos rótulos e campos de entrada
labels = ["Fabrico", "Granulação", "Kwh electrico", "Kwh termico", "Data", "Mail"]
entries = {}

for i, label_text in enumerate(labels):
    label = ttk.Label(root, text=label_text)
    label.grid(row=i, column=0, padx=10, pady=5, sticky="e")
    entry = ttk.Entry(root)
    entry.grid(row=i, column=1, padx=10, pady=5)
    entries[label_text] = entry




# Botão para salvar os dados e enviar SMS
save_and_send_button = ttk.Button(root, text="Salvar e Enviar Email", command=salvar_dados)
save_and_send_button.grid(row=len(labels) + 1, column=0, pady=10)

# Área de exibição dos dados do dia
exibicao_dados = tk.Text(root, height=10, width=40, state=tk.DISABLED)
exibicao_dados.grid(row=len(labels) + 2, columnspan=2, padx=10, pady=10)

# Button to calculate and display the exponential forecast
forecast_exponential_button = ttk.Button(root, text="Previsão Exponencial para o Mês Atual", command=forecast_exponential)
forecast_exponential_button.grid(row=len(labels) + 1, column=1, pady=10)


# Atualiza a exibição com os dados do dia
atualizar_exibicao_dados(10)

# Alteração para aumentar a largura do widget Text
exibicao_dados.configure(width=100)

root.mainloop()
